
from io import BytesIO
import openpyxl


def convert_bytes_like_to_excel_workbook(file_data):
    return openpyxl.load_workbook(BytesIO(file_data))


def extract_workbook_rows(workbook, tab_name=None, max_col=None, max_row=None):
    if not tab_name:
        tab_name = workbook.sheetnames[0]
    ws = workbook[tab_name]
    extracted_rows = []
    for row in ws.iter_rows(min_row=1, max_col=max_col, max_row=max_row):
        row_values = [c.value for c in row]
        extracted_rows.append(row_values)
    return extracted_rows


def extract_excel_rows(file_data):
    workbook = convert_bytes_like_to_excel_workbook(file_data)
    rows = extract_workbook_rows(workbook)
    return rows


def simple_file_type_check(file_data, expected_start=[80, 75, 3, 4, 20, 0]):
    """
    Crude file type check based on finding particular byte sequence at start of files.
    Empirically found to be [80, 75, 3, 4, 20, 0] for Excel files.
    """
    return list(file_data)[:len(expected_start)] == expected_start
